import openpyxl
from openpyxl.styles import PatternFill
import logging
from typing import List, Any

class ColorizedErrors:
    def __init__(self, file_path: str, sheet_name: str, logger):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.logger = logger
        try:
            self.workbook = openpyxl.load_workbook(filename=file_path)
        except FileNotFoundError:
            self.logger.error(f"Файл '{file_path}' не найден.")
            raise
        except Exception as e:
            self.logger.error(f"Ошибка при загрузке файла '{file_path}': {e}")
            raise

    def mark_row(self, search_data: List[Any], color: str) -> None:
        try:
            sheet = self.workbook[self.sheet_name]
        except KeyError:
            self.logger.error(f"Лист '{self.sheet_name}' не найден.")
            return
        
        fill = PatternFill(fill_type="solid", start_color=color, end_color=color)
        search_data_str = [str(val) for val in search_data[:3]]

        found = False
        for row in sheet.iter_rows(min_row=1, max_col=3):
            row_values_str = [str(cell.value) for cell in row]

            if row_values_str == search_data_str:
                row_idx = row[0].row
                self.logger.info(f"Совпадение найдено в строке {row_idx}")

                for cell in sheet[row_idx]:
                    cell.fill = fill
                last_cell = sheet.cell(row=row_idx, column=sheet.max_column)
                last_cell.value = "Удалено"
                
                found = True
                break

        if not found:
            self.logger.warning("Строка для закрашивания не найдена.")
        else:
            try:
                self.workbook.save(self.file_path)
                self.logger.info(f"Строка с данными {search_data_str} найдена и закрашена.")
            except Exception as e:
                self.logger.error(f"Ошибка при сохранении файла '{self.file_path}': {e}")
